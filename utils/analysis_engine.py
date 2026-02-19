import fitz  # PyMuPDF
import openpyxl
import os
import re
import difflib

class AnalysisEngine:
    def __init__(self):
        pass

    def normalize_text(self, text):
        """Normalize text for comparison: lowercase, remove punctuation, collapse spaces."""
        if not text:
            return ""
        text = str(text)
        # Normalize degree symbols (PDF sometimes uses \uf0b0 instead of °)
        text = text.replace("\uf0b0", "°")
        # Replace degree symbol with space so "25°C" and "25 °C" both become "25 c"
        text = text.replace("°", " ")
        for p in [".", ",", ":", ";", "(", ")", "-", "\u2013", "\u2014"]:
            text = text.replace(p, " ")
        return " ".join(text.lower().split())

    def parse_hint(self, hint_text):
        """
        Parse B column hint text to extract search strategies.
        Uses simple sentence-splitting and keyword detection.
        """
        if not hint_text:
            return []
        
        strategies = []
        hint = str(hint_text)
        hint_lower = hint.lower()
        
        # 1. Page hint
        if "ilk sayfa" in hint_lower:
            strategies.append({"type": "page_hint", "page": 0})
        
        # 2. Extract ALL-CAPS phrases as headings (e.g. KISA ÜRÜN BİLGİSİ)
        caps_pattern = r'([A-ZÇĞİÖŞÜ][A-ZÇĞİÖŞÜ\s]{3,}[A-ZÇĞİÖŞÜ])'
        caps_matches = re.findall(caps_pattern, hint)
        for cm in caps_matches:
            heading = cm.strip()
            if len(heading) > 3:
                strategies.append({"type": "under_heading", "heading": heading})
        
        # 3. "X kelimesinden sonra"
        # Filler words to remove from extracted keywords
        filler_words = ["genellikle", "çoğunlukla", "ilk", "sayfada", "sayfasında",
                        "her", "zaman", "olarak", "da", "de", "ve", "ile", "bir",
                        "İlk", "ilk"]  # Turkish İ/i variants
        filler_lower = set(w.casefold() for w in filler_words)
        
        if "kelimesinden sonra" in hint_lower:
            parts = hint.split("kelimesinden")
            if parts:
                before = parts[0].strip()
                for sep in [".", ",", "!"]:
                    if sep in before:
                        before = before.split(sep)[-1].strip()
                keyword = before.rstrip(":").strip()
                # Remove filler words (case-insensitive)
                words = [w for w in keyword.split() if w.casefold() not in filler_lower]
                if len(words) > 3:
                    words = words[-2:]  # Keep last 2 meaningful words
                keyword = " ".join(words)
                if keyword:
                    strategies.append({"type": "after_keyword", "keyword": keyword})
        
        # 4. "X kelimesiyle birlikte"
        if "kelimesiyle birlikte" in hint_lower:
            parts = hint.split("kelimesiyle")
            if parts:
                before = parts[0].strip()
                for sep in [".", ",", "!"]:
                    if sep in before:
                        before = before.split(sep)[-1].strip()
                keyword = before.strip()
                # Remove filler words and percentages
                words = [w for w in keyword.split() if w.casefold() not in filler_lower and not w.startswith("%")]
                if len(words) > 3:
                    words = words[-2:]
                keyword = " ".join(words)
                if keyword:
                    strategies.append({"type": "with_keyword", "keyword": keyword})
        
        # 5. Non-caps headings via "başlığ" pattern
        if "başlığ" in hint_lower or "başlı" in hint_lower:
            for marker in ["başlığının", "başlığı", "başlığ", "başlı"]:
                if marker in hint_lower:
                    idx = hint_lower.find(marker)
                    before = hint[:idx].strip()
                    for sep in [".", ",", "!"]:
                        if sep in before:
                            before = before.split(sep)[-1].strip()
                    for prefix in ["genellikle", "çoğunlukla", "veya", "ya da"]:
                        if before.lower().startswith(prefix):
                            before = before[len(prefix):].strip()
                    if len(before) > 2:
                        strategies.append({"type": "under_heading", "heading": before})
                    break
        
        return strategies

    def contextual_search(self, doc, strategies, ref_name):
        """
        Use parsed strategies to search for content in the PDF.
        Returns: (found, page_num, matched_text) or (False, -1, "")
        """
        page_hints = [s for s in strategies if s["type"] == "page_hint"]
        page_range = [page_hints[0]["page"]] if page_hints else range(len(doc))
        
        for strategy in strategies:
            stype = strategy["type"]
            if stype == "page_hint":
                continue
            
            for page_num in page_range:
                if page_num >= len(doc):
                    continue
                page = doc.load_page(page_num)
                raw_text = page.get_text("text") or ""
                lines = raw_text.split("\n")
                
                if stype == "under_heading":
                    heading = strategy["heading"]
                    heading_norm = self.normalize_text(heading)
                    
                    for i, line in enumerate(lines):
                        line_norm = self.normalize_text(line)
                        if heading_norm in line_norm:
                            context_parts = []
                            for j in range(i+1, min(i+6, len(lines))):
                                stripped = lines[j].strip()
                                if stripped:
                                    context_parts.append(stripped)
                                if len(context_parts) >= 3:
                                    break
                            context_text = " ".join(context_parts)
                            if context_text:
                                return True, page_num, context_text[:200]
                
                elif stype == "after_keyword":
                    keyword = strategy["keyword"]
                    keyword_lower = keyword.lower()
                    
                    for i, line in enumerate(lines):
                        if keyword_lower in line.lower():
                            idx = line.lower().find(keyword_lower)
                            after_text = line[idx + len(keyword):].strip().lstrip(":").strip()
                            if after_text:
                                return True, page_num, after_text[:200]
                            if i + 1 < len(lines) and lines[i+1].strip():
                                return True, page_num, lines[i+1].strip()[:200]
                
                elif stype == "with_keyword":
                    keyword = strategy["keyword"]
                    keyword_norm = self.normalize_text(keyword)
                    page_text_norm = self.normalize_text(raw_text)
                    
                    if keyword_norm in page_text_norm:
                        for i, line in enumerate(lines):
                            if keyword_norm in self.normalize_text(line):
                                return True, page_num, line.strip()[:200]
        
        return False, -1, ""

    def run_analysis(self, rule_excel_path, control_pdf_path):
        """
        Reads Excel (Sheet1) with new 3-column layout:
        A: Reference Name (what to look for)
        B: Hint/Description (where/how to find)
        C: Example Answers (old spellings, separated by '-')
        
        Two-phase search:
        1. Try C column examples directly
        2. If not found, use B column hints for contextual search
        """
        rule_path = rule_excel_path
        control_path = control_pdf_path

        if not os.path.exists(rule_path) or not os.path.exists(control_path):
            return {"status": "Error: File not found.", "results": []}

        try:
            # Load Excel
            try:
                workbook = openpyxl.load_workbook(rule_path)
            except PermissionError:
                return {"status": f"Error: Rule file is open: {os.path.basename(rule_path)}. Please close it.", "results": []}

            if "Sheet1" not in workbook.sheetnames:
                return {"status": "Error: 'Sheet1' not found in Excel file.", "results": []}
            
            sheet = workbook["Sheet1"]
            
            # Load PDF
            try:
                doc = fitz.open(control_pdf_path)
            except PermissionError:
                return {"status": f"Error: Control PDF file is open: {os.path.basename(control_pdf_path)}. Please close it.", "results": []}
            
            # 1. Read Rules from Excel (A, B, C columns)
            rules_data = []
            consecutive_empty = 0
            
            for row in sheet.iter_rows(min_row=2, max_col=3):
                if not row:
                    consecutive_empty += 1
                    if consecutive_empty > 20: break
                    continue
                
                cell_a = row[0] if len(row) > 0 else None
                cell_b = row[1] if len(row) > 1 else None
                cell_c = row[2] if len(row) > 2 else None
                
                ref_name = str(cell_a.value).strip() if cell_a and cell_a.value else ""
                hint = str(cell_b.value).strip() if cell_b and cell_b.value else ""
                examples = str(cell_c.value).strip() if cell_c and cell_c.value else ""
                
                if ref_name:
                    consecutive_empty = 0
                    rules_data.append({
                        "row_index": row[0].row,
                        "ref_name": ref_name,
                        "hint": hint,
                        "examples": examples,
                        "term": examples,  # Backward compat: UI uses "term"
                        "found": False,
                        "matched_term": "",
                        "search_phase": "",  # "examples" or "hint"
                        "locations": []
                    })
                else:
                    consecutive_empty += 1
                    if consecutive_empty > 20: break
            
            if not rules_data:
                return {"status": "No rules found in Excel.", "results": []}

            # Clear debug log
            with open("debug_analysis_engine.txt", "w", encoding="utf-8") as f:
                f.write("=== Analysis Start ===\n")

            # 1.5. PHASE 0: Direct keyword search for specific rules
            # For "Etkin Madde", search for "Etkin madde:" directly
            for rule in rules_data:
                ref_lower = rule["ref_name"].lower()
                if "etkin" not in ref_lower:
                    continue
                
                for page_num in range(len(doc)):
                    if rule["found"]:
                        break
                    page = doc.load_page(page_num)
                    raw_text = page.get_text("text") or ""
                    lines = raw_text.split("\n")
                    
                    for i, line in enumerate(lines):
                        line_lower = line.lower().strip()
                        # Look for "Etkin madde:" or "Etkin maddeler:"
                        if "etkin madde" in line_lower:
                            # Get text after "Etkin madde:" on same line
                            for marker in ["Etkin maddeler:", "Etkin madde:", "etkin maddeler:", "etkin madde:"]:
                                if marker.lower() in line_lower:
                                    idx = line_lower.find(marker.lower())
                                    after = line[idx + len(marker):].strip()
                                    if after:
                                        rule["found"] = True
                                        rule["matched_term"] = after[:200]
                                        rule["search_phase"] = "direct_keyword"
                                        rule["locations"].append({
                                            "page": page_num,
                                            "rect": None,
                                            "matched_term": after[:200]
                                        })
                                        with open("debug_analysis_engine.txt", "a", encoding="utf-8") as f:
                                            f.write(f"[PHASE 0 MATCH] Row {rule['row_index']} '{rule['ref_name']}': Found '{after[:80]}' on page {page_num}\n")
                                        break
                                    else:
                                        # Text might be on next line
                                        if i + 1 < len(lines) and lines[i+1].strip():
                                            next_text = lines[i+1].strip()
                                            rule["found"] = True
                                            rule["matched_term"] = next_text[:200]
                                            rule["search_phase"] = "direct_keyword"
                                            rule["locations"].append({
                                                "page": page_num,
                                                "rect": None,
                                                "matched_term": next_text[:200]
                                            })
                                            with open("debug_analysis_engine.txt", "a", encoding="utf-8") as f:
                                                f.write(f"[PHASE 0 MATCH] Row {rule['row_index']} '{rule['ref_name']}': Found '{next_text[:80]}' on page {page_num}\n")
                                            break
                            if rule["found"]:
                                break

            # 2. PHASE 1: Search C column examples directly
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                raw_text = page.get_text("text") or ""
                for p in [".", ",", ":", ";", "(", ")", "-", "\u2013", "\u2014"]:
                    raw_text = raw_text.replace(p, " ")
                page_text_norm = " ".join(raw_text.lower().split())

                for rule in rules_data:
                    if rule["found"]:
                        continue  # Already matched
                    
                    examples = rule["examples"]
                    if not examples:
                        continue
                    
                    sub_terms = [t.strip() for t in examples.split('-') if t.strip()]
                    
                    for st in sub_terms:
                        st_norm = self.normalize_text(st)
                        if st_norm and st_norm in page_text_norm:
                            rule["found"] = True
                            rule["matched_term"] = st
                            rule["search_phase"] = "examples"
                            rule["locations"].append({
                                "page": page_num,
                                "rect": None,
                                "matched_term": st
                            })
                            with open("debug_analysis_engine.txt", "a", encoding="utf-8") as f:
                                f.write(f"[PHASE 1 MATCH] Row {rule['row_index']} '{rule['ref_name']}': Found '{st}' on page {page_num}\n")
                            break  # Found one sub-term, move to next rule

            # 3. PHASE 2: For unfound rules, use B column hints
            for rule in rules_data:
                if rule["found"]:
                    continue  # Already found in Phase 1
                
                hint = rule["hint"]
                if not hint:
                    continue
                
                strategies = self.parse_hint(hint)
                
                with open("debug_analysis_engine.txt", "a", encoding="utf-8") as f:
                    f.write(f"\n[PHASE 2] Row {rule['row_index']} '{rule['ref_name']}': Strategies={strategies}\n")
                
                if strategies:
                    found, page_num, matched_text = self.contextual_search(doc, strategies, rule["ref_name"])
                    
                    if found:
                        rule["found"] = True
                        rule["matched_term"] = matched_text
                        rule["search_phase"] = "hint"
                        rule["locations"].append({
                            "page": page_num,
                            "rect": None,
                            "matched_term": matched_text
                        })
                        with open("debug_analysis_engine.txt", "a", encoding="utf-8") as f:
                            f.write(f"[PHASE 2 MATCH] Row {rule['row_index']} '{rule['ref_name']}': Found '{matched_text}' on page {page_num}\n")
                    else:
                        with open("debug_analysis_engine.txt", "a", encoding="utf-8") as f:
                            f.write(f"[PHASE 2 FAIL] Row {rule['row_index']} '{rule['ref_name']}': No match\n")

            # 4. PHASE 3: Last resort - scan ALL pages for °C keyword
            # Only for rules that mention °C in their hint (e.g. Saklama Koşulu)
            for rule in rules_data:
                if rule["found"]:
                    continue
                
                hint = rule.get("hint", "")
                if not hint:
                    continue
                
                # Only apply this phase if hint mentions °C
                if "°C" not in hint and "°c" not in hint.lower():
                    continue
                
                with open("debug_analysis_engine.txt", "a", encoding="utf-8") as f:
                    f.write(f"\n[PHASE 3] Row {rule['row_index']} '{rule['ref_name']}': Scanning all pages for °C\n")
                
                for page_num in range(len(doc)):
                    if rule["found"]:
                        break
                    page = doc.load_page(page_num)
                    raw_text = page.get_text("text") or ""
                    # Normalize degree symbols
                    raw_text = raw_text.replace("\uf0b0", "°")
                    lines = raw_text.split("\n")
                    
                    for line in lines:
                        line_stripped = line.strip()
                        if not line_stripped or len(line_stripped) < 5:
                            continue
                        
                        # Only match lines that contain °C
                        if "°C" in line_stripped or "°c" in line_stripped:
                            rule["found"] = True
                            rule["matched_term"] = line_stripped[:200]
                            rule["search_phase"] = "keyword_scan"
                            rule["locations"].append({
                                "page": page_num,
                                "rect": None,
                                "matched_term": line_stripped[:200]
                            })
                            with open("debug_analysis_engine.txt", "a", encoding="utf-8") as f:
                                f.write(f"[PHASE 3 MATCH] Row {rule['row_index']}: Found '{line_stripped[:80]}' on page {page_num}\n")
                            break
                
                if not rule["found"]:
                    with open("debug_analysis_engine.txt", "a", encoding="utf-8") as f:
                        f.write(f"[PHASE 3 FAIL] Row {rule['row_index']} '{rule['ref_name']}': No °C found in PDF\n")

            # 5. Return results
            analysis_results = rules_data
            
        except Exception as e:
            return {"status": f"Error during analysis: {e}", "results": []}

        return {"status": "Analysis Complete", "results": analysis_results}
